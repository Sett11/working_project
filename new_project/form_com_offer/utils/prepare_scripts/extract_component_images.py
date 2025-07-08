#!/usr/bin/env python3
"""
Скрипт для извлечения изображений комплектующих из Excel файла.
"""

import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io

def load_components_catalog(file_path="docs/JSON_files/components_catalog.json"):
    """
    Загружает каталог компонентов из JSON файла.
    
    Args:
        file_path (str): Путь к JSON файлу каталога
        
    Returns:
        dict: Каталог компонентов или None при ошибке
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Файл каталога не найден: {file_path}")
        return None
    except json.JSONDecodeError as e:
        print(f"Ошибка при чтении JSON: {e}")
        return None

def create_images_folder(base_path="docs/images_comp"):
    """
    Создает папку для изображений компонентов.
    
    Args:
        base_path (str): Базовый путь для папки изображений
    """
    # Создаем базовую папку
    if not os.path.exists(base_path):
        os.makedirs(base_path)
        print(f"Создана папка: {base_path}")
    else:
        print(f"Папка уже существует: {base_path}")

def extract_images_from_excel(excel_file_path, output_dir="docs/images_comp"):
    """
    Извлекает изображения из Excel файла.
    
    Args:
        excel_file_path (str): Путь к Excel файлу
        output_dir (str): Папка для сохранения изображений
        
    Returns:
        dict: Словарь с информацией об извлеченных изображениях
    """
    print(f"Извлечение изображений из файла: {excel_file_path}")
    
    try:
        # Загружаем рабочую книгу
        wb = load_workbook(excel_file_path)
        
        extracted_images = {}
        total_images = 0
        
        # Проходим по всем листам
        for sheet_name in wb.sheetnames:
            print(f"Обработка листа: {sheet_name}")
            sheet = wb[sheet_name]
            
            # Получаем все изображения на листе
            if hasattr(sheet, '_images') and sheet._images:
                images = sheet._images
                sheet_images = 0
                
                for img in images:
                    try:
                        # Получаем данные изображения
                        if hasattr(img, 'ref') and hasattr(img.ref, 'blob'):
                            img_data = img.ref.blob
                        elif hasattr(img, 'image'):
                            img_data = img.image.blob
                        else:
                            print(f"  ✗ Неизвестный формат изображения: {type(img)}")
                            continue
                        
                        # Создаем PIL изображение
                        pil_img = Image.open(io.BytesIO(img_data))
                        
                        # Генерируем имя файла
                        image_name = f"{sheet_name}_image_{sheet_images + 1}.jpg"
                        
                        # Сохраняем изображение
                        output_path = os.path.join(output_dir, image_name)
                        pil_img.save(output_path, 'JPEG', quality=90)
                        
                        # Сохраняем информацию
                        extracted_images[image_name] = {
                            'sheet': sheet_name,
                            'path': output_path,
                            'size': pil_img.size,
                            'format': pil_img.format
                        }
                        
                        sheet_images += 1
                        total_images += 1
                        
                        print(f"  ✓ Сохранено: {image_name}")
                        
                    except Exception as e:
                        print(f"  ✗ Ошибка при обработке изображения: {e}")
                        continue
                
                print(f"  Изображений на листе: {sheet_images}")
            else:
                print(f"  На листе {sheet_name} изображений не найдено")
        
        print(f"\nВсего извлечено изображений: {total_images}")
        return extracted_images
        
    except Exception as e:
        print(f"Ошибка при обработке Excel файла: {e}")
        return {}

def read_excel_data(excel_file_path):
    """
    Читает данные из Excel файла для сопоставления с компонентами.
    
    Args:
        excel_file_path (str): Путь к Excel файлу
        
    Returns:
        dict: Данные из Excel файла
    """
    print(f"Чтение данных из Excel файла: {excel_file_path}")
    
    try:
        # Читаем все листы
        excel_data = pd.read_excel(excel_file_path, sheet_name=None)
        
        components_data = {}
        
        for sheet_name, df in excel_data.items():
            print(f"Обработка листа: {sheet_name}")
            print(f"Столбцы: {list(df.columns)}")
            
            # Попытка найти столбцы с названиями компонентов
            name_columns = [col for col in df.columns if isinstance(col, str) and any(keyword in col.lower() for keyword in ['название', 'наименование', 'name', 'компонент', 'материал'])]
            
            if name_columns:
                print(f"  Найдены столбцы с названиями: {name_columns}")
                
                # Берем первые несколько строк для анализа
                sample_data = df.head(20)
                components_data[sheet_name] = {
                    'columns': list(df.columns),
                    'sample_data': sample_data.to_dict('records'),
                    'total_rows': len(df)
                }
            else:
                print(f"  Столбцы с названиями не найдены")
        
        return components_data
        
    except Exception as e:
        print(f"Ошибка при чтении Excel файла: {e}")
        return {}

def match_images_to_components(extracted_images, components_data, catalog):
    """
    Сопоставляет извлеченные изображения с компонентами каталога.
    
    Args:
        extracted_images (dict): Извлеченные изображения
        components_data (dict): Данные из Excel файла
        catalog (dict): Каталог компонентов
        
    Returns:
        dict: Сопоставленные изображения
    """
    print("Сопоставление изображений с компонентами...")
    
    matched_images = {}
    
    # Здесь должна быть логика сопоставления изображений с компонентами
    # Пока что просто возвращаем информацию о найденных изображениях
    for image_name, image_info in extracted_images.items():
        print(f"  Изображение: {image_name}")
        print(f"    Лист: {image_info['sheet']}")
        print(f"    Размер: {image_info['size']}")
        print(f"    Путь: {image_info['path']}")
    
    return matched_images

def main():
    """
    Основная функция.
    """
    print("=" * 60)
    print("Извлечение изображений комплектующих из Excel файла")
    print("=" * 60)
    
    # Путь к Excel файлу
    excel_file_path = "docs/prices_air_and_complectations/стоимости материалов кондиц.xlsx"
    
    # Проверяем существование файла
    if not os.path.exists(excel_file_path):
        print(f"Файл не найден: {excel_file_path}")
        return
    
    # Загружаем каталог компонентов
    catalog = load_components_catalog()
    if not catalog:
        print("Не удалось загрузить каталог компонентов!")
        return
    
    print(f"Загружен каталог с {len(catalog['components'])} компонентами")
    
    # Создаем папку для изображений
    create_images_folder()
    
    # Читаем данные из Excel файла
    components_data = read_excel_data(excel_file_path)
    
    # Извлекаем изображения
    extracted_images = extract_images_from_excel(excel_file_path)
    
    if extracted_images:
        # Сопоставляем изображения с компонентами
        matched_images = match_images_to_components(extracted_images, components_data, catalog)
        
        print(f"\n✓ Обработка завершена!")
        print(f"✓ Извлечено изображений: {len(extracted_images)}")
        print(f"✓ Сопоставлено с компонентами: {len(matched_images)}")
    else:
        print("\n⚠ Изображения не найдены или произошла ошибка при извлечении!")
    
    print("=" * 60)

if __name__ == "__main__":
    main()
